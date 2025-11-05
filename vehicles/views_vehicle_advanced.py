from django.views.generic import View, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from .models import Vehicle, Maintenance, Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import qrcode
from io import BytesIO
import csv
from datetime import datetime

# ============= EXPORTAR A EXCEL =============
class VehicleExportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        # Obtener filtros de la URL
        search = request.GET.get('search', '')
        station = request.GET.get('station', '')
        status = request.GET.get('status', '')
        year = request.GET.get('year', '')
        
        # Aplicar filtros
        vehicles = Vehicle.objects.all()
        if search:
            from django.db.models import Q
            vehicles = vehicles.filter(
                Q(plate__icontains=search) |
                Q(brand__icontains=search) |
                Q(model__icontains=search)
            )
        if station:
            vehicles = vehicles.filter(station=station)
        if status:
            vehicles = vehicles.filter(status=status)
        if year:
            vehicles = vehicles.filter(year=year)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vehículos"
        
        # Estilos
        header_fill = PatternFill(start_color="80AD46", end_color="80AD46", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ['Placa', 'Marca', 'Modelo', 'Año', 'Estado', 'Estación', 'Fecha Registro']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for vehicle in vehicles:
            ws.append([
                vehicle.plate,
                vehicle.brand,
                vehicle.model,
                vehicle.year,
                vehicle.status,
                vehicle.station or 'N/A',
                vehicle.created_at.strftime('%Y-%m-%d') if vehicle.created_at else 'N/A'
            ])
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=vehiculos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


# ============= EXPORTAR A PDF =============
class VehicleExportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        # Obtener filtros
        search = request.GET.get('search', '')
        station = request.GET.get('station', '')
        status = request.GET.get('status', '')
        year = request.GET.get('year', '')
        
        # Aplicar filtros
        vehicles = Vehicle.objects.all()
        if search:
            from django.db.models import Q
            vehicles = vehicles.filter(
                Q(plate__icontains=search) |
                Q(brand__icontains=search) |
                Q(model__icontains=search)
            )
        if station:
            vehicles = vehicles.filter(station=station)
        if status:
            vehicles = vehicles.filter(status=status)
        if year:
            vehicles = vehicles.filter(year=year)
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#80AD46'),
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph("Reporte de Vehículos", title_style))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Tabla
        data = [['Placa', 'Marca', 'Modelo', 'Año', 'Estado', 'Estación']]
        for vehicle in vehicles:
            data.append([
                vehicle.plate,
                vehicle.brand,
                vehicle.model,
                str(vehicle.year),
                vehicle.status,
                vehicle.station or 'N/A'
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#80AD46')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=vehiculos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return response


# ============= IMPORTAR DESDE EXCEL =============
class VehicleImportExcelView(LoginRequiredMixin, View):
    def post(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})
        
        file = request.FILES['file']
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            created = 0
            updated = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Si no hay placa, saltar
                    continue
                
                try:
                    plate, brand, model, year, status, station = row[:6]
                    
                    # Validar año
                    try:
                        year = int(year)
                    except:
                        errors.append(f"Fila {row_num}: Año inválido")
                        continue
                    
                    # Crear o actualizar
                    vehicle, created_flag = Vehicle.objects.update_or_create(
                        plate=plate,
                        defaults={
                            'brand': brand,
                            'model': model,
                            'year': year,
                            'status': status or 'Active',
                            'station': station
                        }
                    )
                    
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
                        
                except Exception as e:
                    errors.append(f"Fila {row_num}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'created': created,
                'updated': updated,
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# ============= HISTORIAL DEL VEHÍCULO =============
class VehicleHistoryView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'vehicles/vehicle_history.html'
    context_object_name = 'vehicle'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehicle = self.object
        
        # Obtener mantenimientos
        maintenances = Maintenance.objects.filter(vehicle=vehicle).order_by('-created_at')
        
        # Obtener documentos
        documents = Document.objects.filter(vehicle=vehicle).order_by('-created_at')
        
        # Crear timeline combinado
        timeline = []
        
        for maint in maintenances:
            timeline.append({
                'type': 'maintenance',
                'date': maint.created_at,
                'title': maint.maintenance_type.name,
                'status': maint.status,
                'description': maint.problem_description,
                'cost': maint.cost,
                'workshop': maint.workshop.name if maint.workshop else 'N/A'
            })
        
        for doc in documents:
            timeline.append({
                'type': 'document',
                'date': doc.created_at,
                'title': doc.doc_type.name,
                'status': 'Expirado' if doc.is_expired else 'Vigente',
                'description': f"Vence: {doc.expiry_date.strftime('%d/%m/%Y')}",
                'number': doc.document_number
            })
        
        # Ordenar por fecha
        timeline.sort(key=lambda x: x['date'], reverse=True)
        
        context['timeline'] = timeline
        return context


# ============= GENERAR QR CODE =============
class VehicleQRCodeView(LoginRequiredMixin, View):
    def get(self, request, pk):
        vehicle = get_object_or_404(Vehicle, pk=pk)
        
        # Datos del QR
        qr_data = f"""Vehículo: {vehicle.plate}
Marca: {vehicle.brand}
Modelo: {vehicle.model}
Año: {vehicle.year}
Estado: {vehicle.status}
Estación: {vehicle.station or 'N/A'}"""
        
        # Generar QR
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="#80AD46", back_color="white")
        
        # Guardar en buffer
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename=qr_{vehicle.plate}.png'
        return response


# ============= DESCARGAR PLANTILLA EXCEL =============
class VehicleDownloadTemplateView(LoginRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Vehículos"
        
        # Estilos
        header_fill = PatternFill(start_color="80AD46", end_color="80AD46", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ['Placa', 'Marca', 'Modelo', 'Año', 'Estado', 'Estación']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ejemplo
        ws.append(['ABC-123', 'Toyota', 'Hilux', '2023', 'Active', 'Ciudad de México'])
        ws.append(['XYZ-789', 'Ford', 'Ranger', '2022', 'Active', 'Jalisco'])
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_vehiculos.xlsx'
        wb.save(response)
        return response
